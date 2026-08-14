from __future__ import annotations

import html
import json
from collections import Counter, defaultdict
from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from examdesk.time_display import (
    DISPLAY_DATETIME_FORMAT,
    EXCEL_DATETIME_FORMAT,
    excel_local_datetime,
    format_local_datetime,
)


class ResultReportService:
    def __init__(self, database) -> None:
        self.database = database

    def export_excel(self, session_id: str, path: Path) -> Path:
        data = self._load_session_data(session_id)
        workbook = Workbook()
        workbook.remove(workbook.active)
        self._summary_sheet(workbook, data)
        self._answers_sheet(workbook, data)
        self._similar_sheet(workbook, data)
        self._focus_sheet(workbook, data)
        self._anomaly_sheet(workbook, data)
        self._session_sheet(workbook, data)
        for worksheet in workbook.worksheets:
            _style_sheet(worksheet)
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(path)
        return path

    def export_candidate_pdf(self, attempt_id: str, path: Path) -> Path:
        with self.database.connect() as connection:
            attempt = connection.execute(
                """
                SELECT a.*, s.name AS session_name FROM attempts a
                LEFT JOIN sessions s ON s.id = a.session_id WHERE a.id = ?
                """,
                (attempt_id,),
            ).fetchone()
            if attempt is None:
                raise KeyError(attempt_id)
            answers = connection.execute(
                "SELECT * FROM attempt_answers WHERE attempt_id = ? ORDER BY display_order",
                (attempt_id,),
            ).fetchall()
            focus_events = connection.execute(
                "SELECT * FROM foreground_events WHERE attempt_id = ? ORDER BY started_at",
                (attempt_id,),
            ).fetchall()
        _register_chinese_font()
        styles = getSampleStyleSheet()
        body = ParagraphStyle(
            "ChineseBody",
            parent=styles["BodyText"],
            fontName="STSong-Light",
            fontSize=9,
            leading=14,
            alignment=TA_LEFT,
            wordWrap="CJK",
        )
        heading = ParagraphStyle(
            "ChineseHeading",
            parent=body,
            fontSize=16,
            leading=22,
            spaceAfter=8,
        )
        small = ParagraphStyle("ChineseSmall", parent=body, fontSize=8, textColor=colors.HexColor("#555555"))
        path.parent.mkdir(parents=True, exist_ok=True)
        document = SimpleDocTemplate(
            str(path),
            pagesize=A4,
            leftMargin=16 * mm,
            rightMargin=16 * mm,
            topMargin=15 * mm,
            bottomMargin=15 * mm,
            title=f"{attempt['candidate_name']}答题记录",
        )
        story = [
            Paragraph(_escape(attempt["session_name"] or "考试答题记录"), heading),
            Table(
                [
                    ["姓名", attempt["candidate_name"], "作答编号", attempt["id"]],
                    ["当前得分", attempt["strict_score"], "预估最高", attempt["estimated_score"]],
                    ["最终得分", attempt["final_score"] or attempt["strict_score"], "满分", attempt["max_score"]],
                    [
                        "开始时间",
                        format_local_datetime(attempt["started_at"], empty="-"),
                        "交卷时间",
                        format_local_datetime(attempt["submitted_at"], empty="-"),
                    ],
                    ["机器", attempt["machine_name"], "Windows用户", attempt["windows_user"]],
                ],
                colWidths=[25 * mm, 52 * mm, 25 * mm, 76 * mm],
                style=_pdf_table_style(),
            ),
            Spacer(1, 8 * mm),
        ]
        for answer in answers:
            snapshot = json.loads(answer["snapshot_json"] or "{}")
            response = _response_text(json.loads(answer["response_json"]))
            correct = _correct_answer_text(snapshot)
            story.extend(
                [
                    Paragraph(f"第{answer['display_order']}题  {_escape(snapshot.get('stem', ''))}", body),
                    Paragraph(f"考生答案：{_escape(response or '未作答')}", body),
                    Paragraph(f"正确答案：{_escape(correct)}", body),
                    Paragraph(
                        "得分：{}/{}　依据：{}".format(
                            answer["final_score"] or answer["strict_score"],
                            _question_max_score(snapshot),
                            _escape(snapshot.get("basis", "")),
                        ),
                        small,
                    ),
                    Spacer(1, 4 * mm),
                ]
            )
        if focus_events:
            story.append(PageBreak())
            story.append(Paragraph("切屏记录", heading))
            focus_rows = [["开始", "时长(秒)", "软件", "窗口标题"]]
            focus_rows.extend(
                [
                    format_local_datetime(event["started_at"], empty="-"),
                    round(event["duration_seconds"], 2),
                    event["process_name"],
                    event["window_title"],
                ]
                for event in focus_events
            )
            story.append(
                Table(
                    focus_rows,
                    repeatRows=1,
                    colWidths=[43 * mm, 22 * mm, 35 * mm, 78 * mm],
                    style=_pdf_table_style(),
                )
            )
        story.append(Spacer(1, 6 * mm))
        story.append(Paragraph(f"核验编号：{attempt['id']}", small))
        document.build(story)
        return path

    def _load_session_data(self, session_id: str) -> dict:
        with self.database.connect() as connection:
            session = connection.execute("SELECT * FROM sessions WHERE id = ?", (session_id,)).fetchone()
            if session is None:
                raise KeyError(session_id)
            attempts = connection.execute(
                "SELECT * FROM attempts WHERE session_id = ? ORDER BY candidate_name, submitted_at",
                (session_id,),
            ).fetchall()
            answers = connection.execute(
                """
                SELECT aa.*, a.candidate_name FROM attempt_answers aa
                JOIN attempts a ON a.id = aa.attempt_id
                WHERE a.session_id = ? ORDER BY a.candidate_name, aa.display_order
                """,
                (session_id,),
            ).fetchall()
            events = connection.execute(
                """
                SELECT fe.*, a.candidate_name FROM foreground_events fe
                JOIN attempts a ON a.id = fe.attempt_id
                WHERE a.session_id = ? ORDER BY a.candidate_name, fe.started_at
                """,
                (session_id,),
            ).fetchall()
            roster = connection.execute(
                "SELECT * FROM session_roster WHERE session_id = ? ORDER BY display_name",
                (session_id,),
            ).fetchall()
            reviews = connection.execute(
                """
                SELECT sr.*, a.candidate_name FROM score_reviews sr
                JOIN attempts a ON a.id = sr.attempt_id
                WHERE a.session_id = ? ORDER BY sr.created_at
                """,
                (session_id,),
            ).fetchall()
        return {
            "session": session,
            "attempts": attempts,
            "answers": answers,
            "events": events,
            "roster": roster,
            "reviews": reviews,
        }

    @staticmethod
    def _summary_sheet(workbook: Workbook, data: dict) -> None:
        worksheet = workbook.create_sheet("成绩汇总")
        worksheet.append(
            [
                "序号", "姓名", "部门", "当前得分", "预估最高", "最终得分", "满分",
                "开始时间", "交卷时间", "用时(分钟)", "交卷方式", "切屏次数", "切屏秒数",
                "状态", "机器", "Windows用户", "作答编号",
            ]
        )
        departments = {row["display_name"].casefold(): row["department"] for row in data["roster"]}
        events_by_attempt = defaultdict(list)
        for event in data["events"]:
            events_by_attempt[event["attempt_id"]].append(event)
        for index, attempt in enumerate(data["attempts"], start=1):
            events = events_by_attempt[attempt["id"]]
            worksheet.append(
                [
                    index,
                    attempt["candidate_name"],
                    departments.get(attempt["candidate_name"].casefold(), ""),
                    _number(attempt["strict_score"]),
                    _number(attempt["estimated_score"]),
                    _number(attempt["final_score"] or attempt["strict_score"]),
                    _number(attempt["max_score"]),
                    excel_local_datetime(attempt["started_at"]),
                    excel_local_datetime(attempt["submitted_at"]),
                    _duration_minutes(attempt["started_at"], attempt["submitted_at"]),
                    attempt["submit_reason"],
                    len(events),
                    round(sum(event["duration_seconds"] or 0 for event in events), 2),
                    "作废" if attempt["is_void"] else attempt["status"],
                    attempt["machine_name"],
                    attempt["windows_user"],
                    attempt["id"],
                ]
            )

    @staticmethod
    def _answers_sheet(workbook: Workbook, data: dict) -> None:
        worksheet = workbook.create_sheet("作答明细")
        worksheet.append(
            ["姓名", "题号", "题目", "考生答案", "正确答案", "当前得分", "预估得分", "最终得分", "依据", "作答编号"]
        )
        for answer in data["answers"]:
            snapshot = json.loads(answer["snapshot_json"] or "{}")
            worksheet.append(
                [
                    answer["candidate_name"],
                    answer["display_order"],
                    snapshot.get("stem", ""),
                    _response_text(json.loads(answer["response_json"])),
                    _correct_answer_text(snapshot),
                    _number(answer["strict_score"]),
                    _number(answer["estimated_score"]),
                    _number(answer["final_score"] or answer["strict_score"]),
                    snapshot.get("basis", ""),
                    answer["attempt_id"],
                ]
            )

    @staticmethod
    def _similar_sheet(workbook: Workbook, data: dict) -> None:
        worksheet = workbook.create_sheet("相似答案复核")
        worksheet.append(
            [
                "姓名", "题号", "空序号", "考生答案", "最相似答案",
                "相似度", "复核结论", "复核后题目得分", "复核人", "说明",
            ]
        )
        latest_reviews = {
            (row["attempt_id"], row["question_id"], row["blank_index"]): row
            for row in data["reviews"]
        }
        for answer in data["answers"]:
            for flag in json.loads(answer["similar_flags_json"]):
                review = latest_reviews.get(
                    (answer["attempt_id"], answer["question_id"], int(flag["blank_index"]))
                )
                worksheet.append(
                    [
                        answer["candidate_name"],
                        answer["display_order"],
                        flag["blank_index"],
                        flag["response"],
                        flag["accepted_answer"],
                        flag["similarity"],
                        review["decision"] if review else "待复核",
                        _number(review["score_after"]) if review else "",
                        review["reviewer_id"] if review else "",
                        review["note"] if review else "",
                    ]
                )

    @staticmethod
    def _focus_sheet(workbook: Workbook, data: dict) -> None:
        worksheet = workbook.create_sheet("切屏记录")
        worksheet.append(["姓名", "开始时间", "结束时间", "持续秒数", "软件", "进程", "窗口标题", "类型", "作答编号"])
        for event in data["events"]:
            worksheet.append(
                [
                    event["candidate_name"],
                    excel_local_datetime(event["started_at"]),
                    excel_local_datetime(event["ended_at"]),
                    round(event["duration_seconds"] or 0, 2), event["application_name"],
                    event["process_name"], event["window_title"], event["event_kind"], event["attempt_id"],
                ]
            )

    @staticmethod
    def _anomaly_sheet(workbook: Workbook, data: dict) -> None:
        worksheet = workbook.create_sheet("异常与未交")
        worksheet.append(["姓名", "异常类型", "说明", "机器", "作答编号"])
        counts = Counter(
            row["candidate_name"].casefold() for row in data["attempts"] if not row["is_void"]
        )
        for attempt in data["attempts"]:
            anomalies = []
            if attempt["is_void"]:
                anomalies.append(("已作废", attempt["void_reason"]))
            if counts[attempt["candidate_name"].casefold()] > data["session"]["max_attempts"]:
                anomalies.append(("重复作答", "超过场次允许次数"))
            if attempt["time_anomaly"]:
                anomalies.append(("时间异常", "检测到系统时间回拨"))
            if attempt["monitor_status"] not in ("ok", "not_started", "disabled"):
                anomalies.append(("监控异常", attempt["monitor_status"]))
            for anomaly_type, detail in anomalies:
                worksheet.append(
                    [attempt["candidate_name"], anomaly_type, detail, attempt["machine_name"], attempt["id"]]
                )
        submitted_names = {row["candidate_name"].casefold() for row in data["attempts"] if not row["is_void"]}
        for roster in data["roster"]:
            if roster["display_name"].casefold() not in submitted_names:
                worksheet.append([roster["display_name"], "未交卷", roster["department"], "", ""])

    @staticmethod
    def _session_sheet(workbook: Workbook, data: dict) -> None:
        worksheet = workbook.create_sheet("场次信息")
        session = data["session"]
        worksheet.append(["项目", "内容"])
        entries = [
            ("场次名称", session["name"]),
            ("场次编号", session["id"]),
            ("考试时长", "不限时" if session["duration_minutes"] is None else f"{session['duration_minutes']}分钟"),
            ("最大答题次数", session["max_attempts"]),
            ("查看策略", session["review_policy"]),
            ("名单人数", len(data["roster"])),
            ("收到记录", len(data["attempts"])),
            ("有效记录", sum(not row["is_void"] for row in data["attempts"])),
            ("题型数量", session["question_counts_json"]),
        ]
        for key, value in entries:
            worksheet.append([key, value])


def _style_sheet(worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="315C4A")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for column_cells in worksheet.columns:
        width = min(
            45,
            max(10, max(_cell_display_length(cell.value) for cell in column_cells) + 2),
        )
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
        for cell in column_cells[1:]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(cell.value, datetime):
                cell.number_format = EXCEL_DATETIME_FORMAT


def _cell_display_length(value) -> int:
    if isinstance(value, datetime):
        return len(value.strftime(DISPLAY_DATETIME_FORMAT))
    return len(str(value or ""))


def _register_chinese_font() -> None:
    if "STSong-Light" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))


def _pdf_table_style() -> TableStyle:
    return TableStyle(
        [
            ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E6EFEA")),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#9AA7A0")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ]
    )


def _correct_answer_text(snapshot: dict) -> str:
    if snapshot.get("question_type") == "fill":
        return "；".join(
            "/".join(str(value) for value in blank.get("accepted_answers", []))
            for blank in snapshot.get("blanks", [])
        )
    return "".join(snapshot.get("correct_option_keys", []))


def _question_max_score(snapshot: dict) -> str:
    return str(snapshot.get("score", ""))


def _response_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return "；".join(str(item) for item in value)
    return str(value)


def _number(value):
    if value in (None, ""):
        return ""
    return float(Decimal(str(value)))


def _duration_minutes(start_value, end_value):
    if not start_value or not end_value:
        return ""
    return round(
        (datetime.fromisoformat(end_value) - datetime.fromisoformat(start_value)).total_seconds() / 60,
        2,
    )


def _escape(value) -> str:
    return html.escape(str(value or "")).replace("\n", "<br/>")
