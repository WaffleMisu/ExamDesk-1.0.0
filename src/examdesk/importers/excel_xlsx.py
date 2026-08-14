from __future__ import annotations

from pathlib import Path, PurePath

from openpyxl import load_workbook

from examdesk.domain.enums import QuestionType
from examdesk.domain.models import QuestionOption
from examdesk.questions.validation import validate_question

from .legacy_txt import OPTION_KEYS, REQUIRED_HEADERS, parse_legacy_row
from .models import ImportCandidate, ImportIssue, ImportPreview, PendingImage

IMAGE_OWNER_BY_HEADER = {
    "题目": "stem",
    "题图": "stem",
    "A": "A",
    "A图": "A",
    "B": "B",
    "B图": "B",
    "C": "C",
    "C图": "C",
    "D": "D",
    "D图": "D",
}


def parse_excel_xlsx(path: Path) -> ImportPreview:
    workbook = load_workbook(path, data_only=False)
    candidates: list[ImportCandidate] = []
    issues: list[ImportIssue] = []
    parsed_sheet_count = 0
    seen_numbers: dict[str, str] = {}

    for worksheet in workbook.worksheets:
        if worksheet.title.strip() in {"字段说明", "使用说明", "说明"}:
            continue
        header_row = _find_header_row(worksheet)
        if header_row is None:
            issues.append(
                ImportIssue(0, "warning", "ignored_sheet", "", f"工作表“{worksheet.title}”没有题库表头")
            )
            continue
        headers = [
            str(worksheet.cell(header_row, column).value or "").strip()
            for column in range(1, worksheet.max_column + 1)
        ]
        missing = REQUIRED_HEADERS - set(headers)
        if missing:
            issues.append(
                ImportIssue(
                    header_row,
                    "error",
                    "missing_headers",
                    "",
                    "工作表“{}”缺少表头：{}".format(worksheet.title, "、".join(sorted(missing))),
                )
            )
            continue
        parsed_sheet_count += 1
        embedded_images = _extract_sheet_images(worksheet, headers)

        for row_number in range(header_row + 1, worksheet.max_row + 1):
            cells = [worksheet.cell(row_number, column) for column in range(1, len(headers) + 1)]
            if not any(cell.value not in (None, "") for cell in cells) and row_number not in embedded_images:
                continue
            row = {header: _cell_text(cell.value) for header, cell in zip(headers, cells, strict=True)}
            for header, cell in zip(headers, cells, strict=True):
                if cell.data_type == "f":
                    issues.append(
                        ImportIssue(
                            row_number,
                            "warning",
                            "formula_cell",
                            header,
                            "公式不会在导入时重新计算，请确认Excel已保存最新结果",
                        )
                    )
            location = f"{worksheet.title}!{row_number}"
            number = row.get("编号", "").strip()
            if number and number in seen_numbers:
                issues.append(
                    ImportIssue(
                        row_number,
                        "error",
                        "duplicate_number",
                        "编号",
                        f"编号与{seen_numbers[number]}重复",
                    )
                )
            elif number:
                seen_numbers[number] = location
            try:
                legacy_record = parse_legacy_row(row, row_number)
            except ValueError as exc:
                issues.append(ImportIssue(row_number, "error", "parse", "", str(exc)))
                continue

            pending_images = list(embedded_images.get(row_number, []))
            pending_images.extend(_read_relative_images(legacy_record.image_paths, path.parent, row_number, issues))
            _attach_pending_placeholders(legacy_record.question, pending_images)
            for validation_issue in validate_question(legacy_record.question):
                issues.append(
                    ImportIssue(
                        row_number,
                        "error",
                        validation_issue.code,
                        validation_issue.field,
                        validation_issue.message,
                    )
                )
            candidates.append(
                ImportCandidate(
                    location,
                    legacy_record.question,
                    pending_images,
                    frozenset(headers),
                )
            )

    if parsed_sheet_count == 0 and not any(issue.code == "missing_headers" for issue in issues):
        raise ValueError("Excel中没有可识别的题库工作表")
    return ImportPreview("excel", candidates, issues)


def _find_header_row(worksheet) -> int | None:
    for row_number in range(1, min(worksheet.max_row, 10) + 1):
        values = {
            str(worksheet.cell(row_number, column).value or "").strip()
            for column in range(1, worksheet.max_column + 1)
        }
        if {"编号", "题型", "题目"} <= values:
            return row_number
    return None


def _extract_sheet_images(worksheet, headers: list[str]) -> dict[int, list[PendingImage]]:
    result: dict[int, list[PendingImage]] = {}
    for index, image in enumerate(getattr(worksheet, "_images", []), start=1):
        anchor = getattr(image, "anchor", None)
        marker = getattr(anchor, "_from", None)
        if marker is None:
            continue
        row_number = marker.row + 1
        column_number = marker.col + 1
        if not 1 <= column_number <= len(headers):
            continue
        owner = IMAGE_OWNER_BY_HEADER.get(headers[column_number - 1], "stem")
        extension = str(getattr(image, "format", "png") or "png").lower()
        result.setdefault(row_number, []).append(
            PendingImage(owner, image._data(), f"excel-{index}.{extension}")
        )
    return result


def _read_relative_images(
    image_paths: dict[str, tuple[str, ...]],
    base_dir: Path,
    row_number: int,
    issues: list[ImportIssue],
) -> list[PendingImage]:
    result = []
    for owner, values in image_paths.items():
        for value in values:
            pure_path = PurePath(value)
            if pure_path.is_absolute() or ".." in pure_path.parts:
                issues.append(
                    ImportIssue(row_number, "error", "unsafe_path", owner, "图片路径必须位于Excel目录内")
                )
                continue
            path = base_dir / Path(*pure_path.parts)
            try:
                data = path.read_bytes()
            except OSError:
                issues.append(
                    ImportIssue(row_number, "warning", "missing_image", owner, f"未找到图片：{value}")
                )
                continue
            result.append(PendingImage(owner, data, path.name))
    return result


def _attach_pending_placeholders(question, images: list[PendingImage]) -> None:
    by_owner: dict[str, list[str]] = {}
    for index, image in enumerate(images, start=1):
        by_owner.setdefault(image.owner_key, []).append(f"pending:{index}")
    question.question_asset_ids = by_owner.get("stem", [])
    options = list(question.options)
    existing_keys = {option.key.strip().upper() for option in options}
    if question.question_type in (
        QuestionType.SINGLE,
        QuestionType.MULTIPLE,
        QuestionType.JUDGE,
    ):
        for key in OPTION_KEYS:
            if key not in existing_keys and by_owner.get(key):
                options.append(QuestionOption(key, "", tuple(by_owner[key])))
                existing_keys.add(key)
    option_order = {key: index for index, key in enumerate(OPTION_KEYS)}
    options.sort(key=lambda option: option_order.get(option.key.strip().upper(), len(option_order)))
    question.options = [
        QuestionOption(
            option.key,
            option.text,
            tuple(dict.fromkeys((*option.asset_ids, *by_owner.get(option.key, [])))),
        )
        for option in options
    ]


def _cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()
