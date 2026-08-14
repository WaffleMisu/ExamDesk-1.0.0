from __future__ import annotations

import os
import sys
import tempfile
from copy import copy
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment
from openpyxl.utils.units import pixels_to_EMU

from examdesk.domain.enums import QuestionType
from examdesk.domain.models import QuestionDraft

WORKBOOK_NAME = "ExamDesk_题库维护模板.xlsx"
HEADERS = (
    "编号",
    "章节",
    "题型",
    "题目",
    "A",
    "B",
    "C",
    "D",
    "答案",
    "依据",
    "分值",
    "题图",
    "A图",
    "B图",
    "C图",
    "D图",
    "状态",
    "使用范围",
    "适用年份",
    "来源",
    "条款",
    "难度",
    "标签",
)
QUESTION_TYPE_TEXT = {
    QuestionType.SINGLE: "单选",
    QuestionType.MULTIPLE: "多选",
    QuestionType.JUDGE: "判断",
    QuestionType.FILL: "填空",
}
IMAGE_COLUMN_BY_OWNER = {"stem": 12, "A": 13, "B": 14, "C": 15, "D": 16}
MAX_IMAGE_WIDTH = 150
MAX_IMAGE_HEIGHT = 110
MAX_IMAGE_STACK_HEIGHT = 520
IMAGE_GAP = 8
IMAGE_PADDING = 6


class ExcelExportError(ValueError):
    pass


@dataclass(frozen=True, slots=True)
class SkippedExportImage:
    display_number: str
    owner_key: str
    asset_id: str
    reason: str


@dataclass(frozen=True, slots=True)
class ExcelExportResult:
    path: Path
    question_count: int
    image_count: int
    skipped_images: tuple[SkippedExportImage, ...]


class QuestionExcelExporter:
    def __init__(self, repository, asset_manager, template_path: Path | None = None) -> None:
        self.repository = repository
        self.asset_manager = asset_manager
        self.template_path = template_path or find_workbook_template()

    def export(self, destination: Path, question_ids=None) -> ExcelExportResult:
        selected = set(question_ids) if question_ids is not None else None
        questions = [
            question
            for question, _version in self.repository.list_current()
            if selected is None or question.id in selected
        ]
        if not questions:
            raise ExcelExportError("题库中没有可导出的题目")
        destination = destination.expanduser().resolve()
        destination.parent.mkdir(parents=True, exist_ok=True)
        skipped: list[SkippedExportImage] = []
        image_count = 0
        temporary_path: Path | None = None
        workbook = None
        try:
            workbook = load_workbook(self.template_path)
            if "题库" not in workbook.sheetnames or "字段说明" not in workbook.sheetnames:
                raise ExcelExportError("题库维护模板缺少“题库”或“字段说明”工作表")
            worksheet = workbook["题库"]
            _prepare_question_sheet(worksheet, len(questions))
            for row_number, question in enumerate(questions, start=2):
                _write_question_row(worksheet, row_number, question)
                added, row_skipped, image_height = self._write_images(
                    worksheet,
                    row_number,
                    question,
                )
                image_count += added
                skipped.extend(row_skipped)
                worksheet.row_dimensions[row_number].height = max(
                    worksheet.row_dimensions[row_number].height or 0,
                    46,
                    image_height * 0.75,
                )
            temporary_path = _temporary_path(destination)
            workbook.save(temporary_path)
            os.replace(temporary_path, destination)
        except ExcelExportError:
            raise
        except Exception as exc:
            raise ExcelExportError(f"无法生成或保存 Excel：{exc}") from exc
        finally:
            if workbook is not None:
                workbook.close()
            if temporary_path is not None:
                temporary_path.unlink(missing_ok=True)
        return ExcelExportResult(destination, len(questions), image_count, tuple(skipped))

    def _write_images(
        self,
        worksheet,
        row_number: int,
        question: QuestionDraft,
    ) -> tuple[int, list[SkippedExportImage], int]:
        owner_assets = {"stem": tuple(question.question_asset_ids)}
        owner_assets.update({option.key.upper(): tuple(option.asset_ids) for option in question.options})
        added = 0
        skipped: list[SkippedExportImage] = []
        maximum_stack_height = 0
        for owner_key, column_number in IMAGE_COLUMN_BY_OWNER.items():
            asset_ids = owner_assets.get(owner_key, ())
            if not asset_ids:
                continue
            image_paths: list[tuple[str, Path]] = []
            for asset_id in asset_ids:
                try:
                    record = self.asset_manager.get(asset_id)
                    path = self.asset_manager.absolute_path(record)
                    if not path.is_file():
                        raise FileNotFoundError(path)
                    image_paths.append((asset_id, path))
                except (KeyError, OSError) as exc:
                    skipped.append(
                        SkippedExportImage(
                            question.display_number,
                            owner_key,
                            asset_id,
                            str(exc),
                        )
                    )
            if not image_paths:
                continue
            per_image_height = min(
                MAX_IMAGE_HEIGHT,
                max(
                    24,
                    (MAX_IMAGE_STACK_HEIGHT - 2 * IMAGE_PADDING - IMAGE_GAP * (len(image_paths) - 1))
                    // len(image_paths),
                ),
            )
            offset = IMAGE_PADDING
            for asset_id, image_path in image_paths:
                try:
                    image = ExcelImage(image_path)
                except (OSError, ValueError) as exc:
                    skipped.append(
                        SkippedExportImage(
                            question.display_number,
                            owner_key,
                            asset_id,
                            str(exc),
                        )
                    )
                    continue
                width, height = _scaled_size(
                    image.width,
                    image.height,
                    MAX_IMAGE_WIDTH,
                    per_image_height,
                )
                image.width = width
                image.height = height
                image.anchor = OneCellAnchor(
                    _from=AnchorMarker(
                        col=column_number - 1,
                        colOff=pixels_to_EMU(IMAGE_PADDING),
                        row=row_number - 1,
                        rowOff=pixels_to_EMU(offset),
                    ),
                    ext=XDRPositiveSize2D(
                        pixels_to_EMU(width),
                        pixels_to_EMU(height),
                    ),
                )
                worksheet.add_image(image)
                added += 1
                offset += height + IMAGE_GAP
            maximum_stack_height = max(maximum_stack_height, offset + IMAGE_PADDING - IMAGE_GAP)
        return added, skipped, maximum_stack_height


def find_workbook_template() -> Path:
    module_path = Path(__file__).resolve()
    candidates = (
        Path(os.environ.get("EXAMDESK_TEMPLATE_DIR", "")) / WORKBOOK_NAME,
        Path(os.path.dirname(os.path.abspath(sys.executable))) / "templates" / WORKBOOK_NAME,
        module_path.parents[3] / "templates" / WORKBOOK_NAME,
        Path.cwd() / "templates" / WORKBOOK_NAME,
    )
    for path in candidates:
        if str(path) and path.is_file():
            return path
    raise ExcelExportError("未找到题库维护模板，请确认软件 templates 文件夹完整")


def _prepare_question_sheet(worksheet, question_count: int) -> None:
    if tuple(worksheet.cell(1, column).value for column in range(1, len(HEADERS) + 1)) != HEADERS:
        raise ExcelExportError("题库维护模板的表头格式不正确")
    worksheet._images.clear()
    target_last_row = max(worksheet.max_row, question_count + 1)
    for row_number in range(2, target_last_row + 1):
        if row_number > worksheet.max_row:
            _copy_row_style(worksheet, 2, row_number)
        for column in range(1, len(HEADERS) + 1):
            worksheet.cell(row_number, column).value = None
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:W{question_count + 1}"
    worksheet.sheet_view.showGridLines = False


def _copy_row_style(worksheet, source_row: int, target_row: int) -> None:
    for column in range(1, len(HEADERS) + 1):
        source = worksheet.cell(source_row, column)
        target = worksheet.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.protection = copy(source.protection)
        target.alignment = copy(source.alignment)


def _write_question_row(worksheet, row_number: int, question: QuestionDraft) -> None:
    option_text = {option.key.upper(): option.text for option in question.options}
    answer, score = _answer_and_score(question)
    values = (
        question.display_number,
        question.chapter,
        QUESTION_TYPE_TEXT[question.question_type],
        question.stem,
        option_text.get("A", ""),
        option_text.get("B", ""),
        option_text.get("C", ""),
        option_text.get("D", ""),
        answer,
        question.basis,
        score,
        "",
        "",
        "",
        "",
        "",
        _status_text(question.status),
        _scope_text(question.usage_scope),
        question.applicable_year or "",
        question.source,
        question.clause,
        question.difficulty,
        ";".join(question.tags),
    )
    for column, value in enumerate(values, start=1):
        cell = worksheet.cell(row_number, column, value)
        cell.alignment = copy(cell.alignment)
        cell.alignment = Alignment(
            horizontal=cell.alignment.horizontal or "left",
            vertical="top",
            wrap_text=True,
        )


def _answer_and_score(question: QuestionDraft) -> tuple[str, str]:
    if question.question_type is not QuestionType.FILL:
        return "".join(sorted(question.correct_option_keys)), _decimal_text(question.score)
    answers = ";".join(
        "/".join(answer.strip() for answer in blank.accepted_answers if answer.strip())
        for blank in question.blanks
    )
    if question.unordered_groups:
        ranges = []
        for group in question.unordered_groups:
            indexes = tuple(sorted(set(group.indexes)))
            if indexes != tuple(range(indexes[0], indexes[-1] + 1)):
                raise ExcelExportError(
                    f"题号 {question.display_number or question.id} 的无序填空组不是连续范围，无法按原格式导出"
                )
            ranges.append(f"{indexes[0]}-{indexes[-1]}")
        answers = "@{}|{}".format(",".join(ranges), answers)
    scores = [_decimal_text(blank.score) for blank in question.blanks]
    score_text = scores[0] if scores and len(set(scores)) == 1 else ";".join(scores)
    return answers, score_text


def _status_text(status) -> str:
    return {
        "draft": "草稿",
        "enabled": "启用",
        "disabled": "停用",
    }[status.value]


def _scope_text(scope) -> str:
    return {
        "both": "考试和练习",
        "exam_only": "仅考试",
        "practice_only": "仅练习",
    }[scope.value]


def _decimal_text(value: Decimal) -> str:
    normalized = value.normalize()
    return format(normalized, "f")


def _scaled_size(width: int, height: int, maximum_width: int, maximum_height: int) -> tuple[int, int]:
    if width <= 0 or height <= 0:
        raise ExcelExportError("图片尺寸无效")
    scale = min(1.0, maximum_width / width, maximum_height / height)
    return max(1, round(width * scale)), max(1, round(height * scale))


def _temporary_path(destination: Path) -> Path:
    handle, name = tempfile.mkstemp(
        prefix=f".{destination.stem}-",
        suffix=destination.suffix or ".xlsx",
        dir=destination.parent,
    )
    os.close(handle)
    return Path(name)
