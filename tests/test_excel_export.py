from __future__ import annotations

import io
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image

from examdesk.db import Database, initialize_database
from examdesk.domain.enums import QuestionStatus, QuestionType, UsageScope
from examdesk.domain.models import (
    BlankDefinition,
    QuestionDraft,
    QuestionOption,
    UnorderedGroup,
)
from examdesk.importers import parse_excel_xlsx
from examdesk.questions import AssetManager, QuestionExcelExporter, QuestionRepository


def _png_bytes(color: tuple[int, int, int]) -> bytes:
    output = io.BytesIO()
    Image.new("RGB", (180, 100), color).save(output, format="PNG")
    return output.getvalue()


def _services(tmp_path: Path):
    database_path = tmp_path / "data.sqlite3"
    initialize_database(database_path)
    database = Database(database_path)
    return QuestionRepository(database), AssetManager(database, tmp_path / "assets")


def _template_path() -> Path:
    return Path(__file__).parents[1] / "templates" / "ExamDesk_题库维护模板.xlsx"


def test_excel_export_round_trips_choice_fill_and_embedded_images(tmp_path: Path) -> None:
    repository, assets = _services(tmp_path)
    red = assets.ingest_bytes(_png_bytes((190, 45, 55)), "red.png")
    blue = assets.ingest_bytes(_png_bytes((45, 95, 190)), "blue.png")
    green = assets.ingest_bytes(_png_bytes((35, 145, 85)), "green.png")
    choice = QuestionDraft(
        question_type=QuestionType.MULTIPLE,
        stem="根据题图选择正确说法",
        basis="",
        display_number="001",
        chapter="操作认定",
        clause="第三条",
        source="年度培训手册",
        difficulty="较难",
        tags=["耕地", "重点"],
        usage_scope=UsageScope.EXAM_ONLY,
        applicable_year=2026,
        status=QuestionStatus.ENABLED,
        options=[
            QuestionOption("A", "说法甲", (green.id,)),
            QuestionOption("B", "说法乙"),
            QuestionOption("C", "说法丙"),
            QuestionOption("D", "说法丁"),
        ],
        correct_option_keys={"A", "C"},
        question_asset_ids=[red.id, blue.id],
        score=Decimal("2.5"),
    )
    fill = QuestionDraft(
        question_type=QuestionType.FILL,
        stem="填写（1）（2）（3）（4）",
        basis="培训手册",
        display_number="002",
        chapter="填空",
        status=QuestionStatus.ENABLED,
        blanks=[
            BlankDefinition(1, ("甲", "第一"), Decimal("0.5")),
            BlankDefinition(2, ("乙",), Decimal("0.5")),
            BlankDefinition(3, ("丙",), Decimal("1")),
            BlankDefinition(4, ("丁",), Decimal("1")),
        ],
        unordered_groups=[UnorderedGroup((1, 2)), UnorderedGroup((3, 4))],
        score=Decimal("3"),
    )
    repository.create(choice, actor_id=None)
    repository.create(fill, actor_id=None)

    destination = tmp_path / "导出题库.xlsx"
    result = QuestionExcelExporter(repository, assets, _template_path()).export(destination)

    assert result.question_count == 2
    assert result.image_count == 3
    assert result.skipped_images == ()
    workbook = load_workbook(destination)
    sheet = workbook["题库"]
    assert sheet["I2"].value == "AC"
    assert sheet["K2"].value == "2.5"
    assert [sheet.cell(2, column).value for column in range(17, 24)] == [
        "启用", "仅考试", 2026, "年度培训手册", "第三条", "较难", "耕地;重点"
    ]
    assert sheet["I3"].value == "@1-2,3-4|甲/第一;乙;丙;丁"
    assert sheet["K3"].value == "0.5;0.5;1;1"
    anchors = [(image.anchor._from.row, image.anchor._from.col) for image in sheet._images]
    assert anchors.count((1, 11)) == 2
    assert anchors.count((1, 12)) == 1
    workbook.close()

    preview = parse_excel_xlsx(destination)
    assert preview.error_count == 0
    assert len(preview.candidates) == 2
    imported_choice = preview.candidates[0]
    imported_fill = preview.candidates[1].question
    assert imported_choice.question.correct_option_keys == {"A", "C"}
    assert [image.owner_key for image in imported_choice.images].count("stem") == 2
    assert [image.owner_key for image in imported_choice.images].count("A") == 1
    assert imported_fill.unordered_groups == [UnorderedGroup((1, 2)), UnorderedGroup((3, 4))]
    assert [blank.score for blank in imported_fill.blanks] == [
        Decimal("0.5"),
        Decimal("0.5"),
        Decimal("1"),
        Decimal("1"),
    ]


def test_excel_export_uses_one_score_when_every_fill_blank_matches(tmp_path: Path) -> None:
    repository, assets = _services(tmp_path)
    question = QuestionDraft(
        question_type=QuestionType.FILL,
        stem="填写（1）（2）（3）",
        basis="",
        display_number="003",
        status=QuestionStatus.ENABLED,
        blanks=[
            BlankDefinition(1, ("甲",), Decimal("0.25")),
            BlankDefinition(2, ("乙",), Decimal("0.25")),
            BlankDefinition(3, ("丙",), Decimal("0.25")),
        ],
        score=Decimal("0.75"),
    )
    repository.create(question, actor_id=None)

    destination = tmp_path / "同分填空.xlsx"
    QuestionExcelExporter(repository, assets, _template_path()).export(destination)

    workbook = load_workbook(destination)
    assert workbook["题库"]["K2"].value == "0.25"
    workbook.close()


def test_excel_export_reports_missing_asset_without_losing_question(tmp_path: Path) -> None:
    repository, assets = _services(tmp_path)
    image = assets.ingest_bytes(_png_bytes((120, 120, 120)), "missing.png")
    question = QuestionDraft(
        question_type=QuestionType.SINGLE,
        stem="缺图仍应导出",
        basis="",
        display_number="004",
        status=QuestionStatus.ENABLED,
        options=[QuestionOption("A", "正确"), QuestionOption("B", "错误")],
        correct_option_keys={"A"},
        question_asset_ids=[image.id],
        score=Decimal("1"),
    )
    repository.create(question, actor_id=None)
    assets.absolute_path(image).unlink()

    destination = tmp_path / "缺图.xlsx"
    result = QuestionExcelExporter(repository, assets, _template_path()).export(destination)

    assert destination.is_file()
    assert result.image_count == 0
    assert len(result.skipped_images) == 1
    assert result.skipped_images[0].display_number == "004"
    workbook = load_workbook(destination)
    assert workbook["题库"]["D2"].value == "缺图仍应导出"
    workbook.close()
