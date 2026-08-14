import json
from datetime import UTC, datetime, timedelta
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook
from test_exam_runtime import exam_definition
from test_results_flow import create_master_session

from examdesk.db import Database, initialize_database
from examdesk.questions import question_to_payload
from examdesk.results import ResultReportService
from examdesk.time_display import EXCEL_DATETIME_FORMAT


def prepare_report_database(tmp_path: Path):
    database_path = tmp_path / "report.sqlite3"
    initialize_database(database_path)
    database = Database(database_path)
    definition = exam_definition()
    create_master_session(database, definition)
    attempt_id = "attempt-report"
    started = datetime(2026, 8, 4, 9, 0, tzinfo=UTC)
    submitted = started + timedelta(minutes=20)
    with database.transaction(immediate=True) as connection:
        connection.execute(
            """
            INSERT INTO session_roster(session_id, display_name, department, extra_json)
            VALUES (?, '测试用户甲', '第一组', '{}'), (?, '未交人员', '第二组', '{}')
            """,
            (definition.session_id, definition.session_id),
        )
        connection.execute(
            """
            INSERT INTO attempts(
                id, session_id, candidate_name, machine_name, windows_user,
                software_version, status, started_at, deadline_at, submitted_at,
                submit_reason, strict_score, estimated_score, final_score, max_score,
                question_order_json, monitor_status, created_at
            ) VALUES (?, ?, '测试用户甲', 'PC-01', 'test_user', '2.0.0', 'submitted',
                      ?, ?, ?, 'manual', '6', '8', '8', '8', ?, 'ok', ?)
            """,
            (
                attempt_id,
                definition.session_id,
                started.isoformat(),
                (started + timedelta(minutes=30)).isoformat(),
                submitted.isoformat(),
                json.dumps([item.question_id for item in definition.questions]),
                started.isoformat(),
            ),
        )
        for order, item in enumerate(definition.questions, start=1):
            snapshot = question_to_payload(item.question)
            connection.execute(
                """
                INSERT INTO attempt_answers(
                    attempt_id, question_id, display_order, option_order_json,
                    response_json, strict_score, estimated_score, final_score,
                    similar_flags_json, answered_at, snapshot_json
                ) VALUES (?, ?, ?, '[]', ?, '2', '2', '2', '[]', ?, ?)
                """,
                (
                    attempt_id,
                    item.question_id,
                    order,
                    json.dumps(["A"], ensure_ascii=False),
                    submitted.isoformat(),
                    json.dumps(snapshot, ensure_ascii=False),
                ),
            )
        connection.execute(
            """
            INSERT INTO foreground_events(
                id, attempt_id, started_at, ended_at, duration_seconds,
                application_name, process_name, window_title, event_kind, created_at
            ) VALUES (?, ?, ?, ?, 4, '通讯工具', '通讯工具.exe', '文件传输', 'window', ?)
            """,
            (
                str(uuid4()),
                attempt_id,
                (started + timedelta(minutes=2)).isoformat(),
                (started + timedelta(minutes=2, seconds=4)).isoformat(),
                submitted.isoformat(),
            ),
        )
    return database, definition, attempt_id


def test_excel_and_pdf_reports_include_expected_sections(tmp_path: Path) -> None:
    database, definition, attempt_id = prepare_report_database(tmp_path)
    service = ResultReportService(database)
    excel_path = service.export_excel(definition.session_id, tmp_path / "成绩汇总.xlsx")
    pdf_path = service.export_candidate_pdf(attempt_id, tmp_path / "测试用户甲答卷.pdf")

    workbook = load_workbook(excel_path)
    assert workbook.sheetnames == [
        "成绩汇总",
        "作答明细",
        "相似答案复核",
        "切屏记录",
        "异常与未交",
        "场次信息",
    ]
    assert workbook["成绩汇总"]["B2"].value == "测试用户甲"
    expected_started = datetime(2026, 8, 4, 9, 0, tzinfo=UTC).astimezone().replace(
        tzinfo=None,
        microsecond=0,
    )
    assert workbook["成绩汇总"]["H2"].value == expected_started
    assert workbook["成绩汇总"]["H2"].number_format == EXCEL_DATETIME_FORMAT
    assert workbook["成绩汇总"]["I2"].number_format == EXCEL_DATETIME_FORMAT
    assert workbook["切屏记录"]["B2"].value == expected_started + timedelta(minutes=2)
    assert workbook["切屏记录"]["B2"].number_format == EXCEL_DATETIME_FORMAT
    assert workbook["切屏记录"]["C2"].number_format == EXCEL_DATETIME_FORMAT
    anomaly_values = [cell.value for row in workbook["异常与未交"].iter_rows() for cell in row]
    assert "未交人员" in anomaly_values
    assert pdf_path.read_bytes().startswith(b"%PDF")
    assert pdf_path.stat().st_size > 2000
    workbook.close()
