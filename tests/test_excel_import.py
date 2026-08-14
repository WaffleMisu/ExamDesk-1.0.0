from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image

from examdesk.domain.enums import QuestionStatus, QuestionType, UsageScope
from examdesk.importers import parse_excel_xlsx

HEADERS = [
    "编号", "章节", "题型", "题目", "A", "B", "C", "D",
    "答案", "依据", "分值", "题图", "A图", "B图", "C图", "D图",
]


def make_png(path: Path) -> None:
    Image.new("RGB", (60, 40), (30, 120, 180)).save(path, format="PNG")


def test_excel_import_reads_choice_and_embedded_question_image(tmp_path: Path) -> None:
    image_path = tmp_path / "题图.png"
    make_png(image_path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "题库"
    sheet.append(HEADERS)
    sheet.append(["001", "安全规范", "单选", "应如何认定？", "甲", "乙", "丙", "丁", "B", "培训手册", 2])
    sheet.add_image(ExcelImage(image_path), "L2")
    path = tmp_path / "题库.xlsx"
    workbook.save(path)

    preview = parse_excel_xlsx(path)
    candidate = preview.candidates[0]

    assert preview.error_count == 0
    assert candidate.question.question_type is QuestionType.SINGLE
    assert candidate.question.correct_option_keys == {"B"}
    assert len(candidate.images) == 1
    assert candidate.images[0].owner_key == "stem"


def test_excel_import_reads_three_images_anchored_to_one_question_cell(tmp_path: Path) -> None:
    image_paths = []
    for index, color in enumerate(((180, 40, 40), (40, 150, 70), (40, 90, 180)), start=1):
        image_path = tmp_path / f"题图{index}.png"
        Image.new("RGB", (60, 40), color).save(image_path, format="PNG")
        image_paths.append(image_path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(HEADERS)
    sheet.append(["003", "影像判读", "单选", "根据三张题图判断", "甲", "乙", "", "", "A", "依据", 1])
    for image_path in image_paths:
        sheet.add_image(ExcelImage(image_path), "L2")
    path = tmp_path / "三张题图.xlsx"
    workbook.save(path)

    preview = parse_excel_xlsx(path)

    assert preview.error_count == 0
    assert len(preview.candidates[0].images) == 3
    assert {image.owner_key for image in preview.candidates[0].images} == {"stem"}


def test_excel_import_accepts_image_only_options_in_a_and_b_cells(tmp_path: Path) -> None:
    option_a = tmp_path / "选项A.png"
    option_b = tmp_path / "选项B.png"
    make_png(option_a)
    make_png(option_b)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(HEADERS)
    sheet.append(["020", "影像判读", "单选", "请选择正确图形", "", "", "", "", "A", "", 1])
    sheet.add_image(ExcelImage(option_a), "E2")
    sheet.add_image(ExcelImage(option_b), "F2")
    path = tmp_path / "纯图片选项.xlsx"
    workbook.save(path)

    preview = parse_excel_xlsx(path)
    question = preview.candidates[0].question

    assert preview.error_count == 0
    assert [option.key for option in question.options] == ["A", "B"]
    assert all(option.text == "" for option in question.options)
    assert all(option.asset_ids for option in question.options)
    assert [image.owner_key for image in preview.candidates[0].images] == ["A", "B"]


def test_excel_import_accepts_image_only_options_in_a_image_and_b_image_cells(
    tmp_path: Path,
) -> None:
    option_a = tmp_path / "A图.png"
    option_b = tmp_path / "B图.png"
    make_png(option_a)
    make_png(option_b)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(HEADERS)
    sheet.append(["021", "影像判读", "单选", "请选择正确图形", "", "", "", "", "B", "", 1])
    sheet.add_image(ExcelImage(option_a), "M2")
    sheet.add_image(ExcelImage(option_b), "N2")
    path = tmp_path / "纯图片选项图列.xlsx"
    workbook.save(path)

    preview = parse_excel_xlsx(path)

    assert preview.error_count == 0
    assert [option.key for option in preview.candidates[0].question.options] == ["A", "B"]


def test_excel_import_rejects_choice_with_only_one_image_option(tmp_path: Path) -> None:
    option_a = tmp_path / "唯一选项.png"
    make_png(option_a)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(HEADERS)
    sheet.append(["022", "影像判读", "单选", "请选择正确图形", "", "", "", "", "A", "", 1])
    sheet.add_image(ExcelImage(option_a), "E2")
    path = tmp_path / "单个图片选项.xlsx"
    workbook.save(path)

    preview = parse_excel_xlsx(path)

    assert any(issue.code == "count" and issue.field == "options" for issue in preview.issues)


def test_excel_import_reads_fill_with_single_repeated_score(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(HEADERS)
    sheet.append([
        "002",
        "填空",
        "填空",
        "（1）（2）（3）",
        "",
        "",
        "",
        "",
        "@1-3|甲;乙;丙",
        "培训手册",
        "0.5",
    ])
    path = tmp_path / "填空.xlsx"
    workbook.save(path)

    preview = parse_excel_xlsx(path)
    question = preview.candidates[0].question

    assert preview.error_count == 0
    assert question.question_type is QuestionType.FILL
    assert str(question.score) == "1.5"


def test_excel_import_marks_duplicate_number_across_sheets(tmp_path: Path) -> None:
    workbook = Workbook()
    first = workbook.active
    first.title = "第一章"
    second = workbook.create_sheet("第二章")
    for sheet in (first, second):
        sheet.append(HEADERS)
        sheet.append(["001", "章节", "单选", "题目", "甲", "乙", "", "", "A", "依据", 1])
    path = tmp_path / "重复.xlsx"
    workbook.save(path)

    preview = parse_excel_xlsx(path)
    assert any(issue.code == "duplicate_number" for issue in preview.issues)


def test_excel_import_silently_ignores_field_guide_sheet(tmp_path: Path) -> None:
    workbook = Workbook()
    bank = workbook.active
    bank.title = "题库"
    bank.append(HEADERS)
    bank.append(["004", "章节", "判断", "说法是否正确", "正确", "错误", "", "", "A", "依据", 1])
    guide = workbook.create_sheet("字段说明")
    guide.append(["字段", "填写方式"])
    path = tmp_path / "带说明的模板.xlsx"
    workbook.save(path)

    preview = parse_excel_xlsx(path)

    assert len(preview.candidates) == 1
    assert preview.warning_count == 0


def test_excel_import_reads_optional_management_columns(tmp_path: Path) -> None:
    headers = HEADERS + ["状态", "使用范围", "适用年份", "来源", "条款", "难度", "标签"]
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    sheet.append(
        [
            "010", "年度变更", "单选", "题目", "甲", "乙", "", "", "A", "", 1,
            "", "", "", "", "", "停用", "仅练习", 2026, "培训手册", "第三条", "较难", "耕地;重点;耕地",
        ]
    )
    path = tmp_path / "扩展列.xlsx"
    workbook.save(path)

    preview = parse_excel_xlsx(path)
    loaded = preview.candidates[0].question

    assert preview.error_count == 0
    assert loaded.status is QuestionStatus.DISABLED
    assert loaded.usage_scope is UsageScope.PRACTICE_ONLY
    assert loaded.applicable_year == 2026
    assert (loaded.source, loaded.clause, loaded.difficulty) == ("培训手册", "第三条", "较难")
    assert loaded.tags == ["耕地", "重点"]
